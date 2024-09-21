import requests
from bs4 import BeautifulSoup
import re
import time

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
import pandas as pd
from pandas import concat

#from press release website, scrap stock press releases for the day
def acquire_news_tickers(date):
    data = []
    date.replace("-","%2f")
    
    for num in range(1,2): #adjust the range to adjust the number of news pages you want to scrap, shouldn't be more than the number of pages the website stockhouse.com actually has
        url = f'https://stockhouse.com/news/us-press-releases?sort=DESC&sdate={date}&page={num}'
    
        try:
            response = requests.get(url)
            if response.status_code == 200:
                # Parse raw HTML code and extract relevant sections using BeautifulSoup
                soup = BeautifulSoup(response.text, 'html.parser')
                rows = soup.find_all('td', class_='pr-headline')

                #further extract news titles and company tickers using string methods
                for row in rows:
                    new_row = []
                    row_str = str(row)
                    start_delimiter_1 = r'<a href="'
                    end_delimiter_1 = r'"'
                    start_delimiter_2 = r'target="_blank">'
                    end_delimiter_2 = r'</a>'
                    start_delimiter_3 = r'/companies/quote?symbol='
                    end_delimiter_3 = r'"'

                    result = re.findall(f"{re.escape(start_delimiter_1)}(.*?){re.escape(end_delimiter_1)}", row_str)
                    result2 = re.findall(f"{re.escape(start_delimiter_2)}(.*?){re.escape(end_delimiter_2)}", row_str)
                    result3 = re.findall(f"{re.escape(start_delimiter_3)}(.*?){re.escape(end_delimiter_3)}", row_str)

                    #tidy data and form dataset
                    new_row.append(result2[0])
                    new_row.append(result3[0])
                    full_link = "stockhouse.com" + result[0]
                    new_row.append("stockhouse.com"+result[0])
                    data.append(new_row)
            else:
                print("Failed to connect to servers")
    
        except:
            print("Failed to find url or the website does not have that many pages (reduce the range in this for-loop if that's the case)")
        
    return data

#from api, acquire stock data based on scraped companies
def acquire_stock_movement(ticker,day,key):

    day = day.replace("-2024", "")
    day = "2024-" + day

    response = requests.get(f"https://api.twelvedata.com/time_series?apikey={key}&interval=1day&symbol={ticker}&dp=2&country=US&start_date={day} 00:00:00&end_date={day} 23:59:00")
    response = response.json()

    #calculate stock data
    open_price = float(response["values"][0]["open"])
    close_price = float(response["values"][0]["close"])
    percentage_change = (close_price-open_price)/open_price

    return percentage_change

#main function, acquire all the data and write it into a csv
def get_and_write_all_stock_movements(date,user_key):

    #scrap all data and fetch relevant information from api
    news_data = acquire_news_tickers(date)
    for news in news_data:
        percentage_change = acquire_stock_movement(news[1],date,user_key)
        news.insert(2,f"{percentage_change:2%}")
        print(news[1],news[2])
        time.sleep(8)

    #write everything into a csv
    workbook = Workbook()
    sheet = workbook.active
    sheet['A1'] = 'Company Ticker'
    sheet['B1'] = 'Date'
    sheet['C1'] = 'Stock Percentage Change'
    sheet['D1'] = 'News Link'

    for index, (title, ticker, percentage, url) in enumerate(news_data, start=2):
        sheet.cell(row=index, column=1).value = ticker
        sheet.cell(row=index, column=2).value = date
        sheet.cell(row=index,column=3).value = percentage
        link_cell = sheet.cell(row=index, column=4)
        link_cell.hyperlink = "https://" + url
        link_cell.value = title
        link_cell.style = 'Hyperlink'
        link_cell.font = Font(color='0000FF', underline='single')

    filename = 'NewsLinks.xlsx'
    workbook.save(filename)
    df = pd.read_excel(filename)
    csv_file = "Newslinks.csv"
    df.to_csv(csv_file,index=False)
    print(f"File saved as {filename}")

#feature
your_api_key = ""
user_date = input("Enter a date (MM-DD-YYYY) to check all stock performance and news on that day (input yesterday or any past date)")
get_and_write_all_stock_movements(user_date,your_api_key)




